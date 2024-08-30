import requests  
import json  
from base64 import b64encode
import urllib3
from datetime import datetime 
import pandas as pd
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import pytz
from openpyxl import Workbook, load_workbook

# Suppress the SSL warning
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# Configuration  
WAZUH_API_IP = "https://192.168.129.118:55000"  
USERNAME = "wazuh-wui"  
PASSWORD = "6Ir*J.*xShvkVB1fqQXJ4Fwk9VKNfOpH"
TOKEN = None 

def authenticate():  
    """Authenticate with Wazuh API and return the JWT token."""  
    url = f"{WAZUH_API_IP}/security/user/authenticate"  
    basic_auth = f"{USERNAME}:{PASSWORD}".encode() 
   
    login_headers = {
    'Content-Type': "Application/json",
    'Authorization': f'Basic {b64encode(basic_auth).decode()}'
    }  
    
    response = requests.post(url, headers=login_headers,verify=False)  
    print(response) 
    if response.status_code == 200:  
        token = response.json()['data']['token']  
        print("Authentication successful.")  
        return token  
    else:  
        print("Authentication failed:", response.json())  
        return None  


def print_time_ago(date_time_obj):
    now = datetime.now(pytz.utc)
    
    # Calculate the difference between now and the provided datetime
    delta = relativedelta(now, date_time_obj)
    
    # Print years, months, and days ago
    if delta.years > 0:
        return f"{delta.years} year{'s' if delta.years > 1 else ''} ago"
    elif delta.months > 0:
        return f"{delta.months} month{'s' if delta.months > 1 else ''} ago"
    elif delta.days > 0:
        return f"{delta.days} day{'s' if delta.days > 1 else ''} ago"
    elif delta.hours > 0:
        return f"{delta.hours} hour{'s' if delta.hours > 1 else ''} ago"
    elif delta.minutes > 0:
        return f"{delta.minutes} minute{'s' if delta.minutes > 1 else ''} ago"
    
    elif delta.seconds > 0:
        return f"{delta.seconds} second{'s' if delta.seconds > 1 else ''} ago"
        


def getMoreData(agent_id,token):
    data = {}
    headers = {  
        'Authorization': f'Bearer {token}',  
        'Content-Type': 'application/json'  
    }  

    for path in ['hardware','os']:
        url = f"{WAZUH_API_IP}/syscollector/{agent_id}/{path}/?pretty=true" 
     
        response = requests.get(url, headers=headers,verify=False) 
        if response.status_code == 200 and path == 'hardware':  
            agent = response.json() 
            print(agent)
            agent = agent['data']['affected_items'][0]
            # print(agent)
            if 'cpu' in agent:
                data['processor'] =  f"{agent['cpu']['name']}, cores: {agent['cpu']['cores']}"
            else:
                data['processor'] = "Unknown"

            if 'ram' in agent:
                data['AvailableRAM'] =  f"{agent['ram']['free'] / (1024):.2f} MB / {agent['ram']['total'] / (1024):.2f} MB, Usage: {agent['ram']['usage']}%"
            else:
                data['AvailableRAM'] = "Unknown"    

            # data['AvailableRAM'] = f"{agent['ram']['free'] / (1024):.2f} MB / {agent['ram']['total'] / (1024):.2f} MB, Usage: {agent['ram']['usage']}%"
            data['SerialNumber'] = agent.get('board_serial',"Unknown")
            if 'scan' in agent:
                data['LastScan'] = print_time_ago(datetime.strptime(agent['scan']['time'],"%Y-%m-%dT%H:%M:%S+00:00").replace(tzinfo=pytz.utc))

            else:
                data['LastScan'] = "Unknown"

        elif response.status_code == 200 and path == 'os':
              agent = response.json() 
              

    return data 

def saveToExcel(index, data):
    
   existing_exel = load_workbook("./agent.xlsx")
   sheet = existing_exel.active 
   rows = sheet.max_row + 1
   row_data = [rows] + [data[key] for key in data]
   print(row_data)
   for i in range(len(row_data)):
        sheet.cell(row=rows,column=i+1).value = row_data[i] 
        existing_exel.save("./agent.xlsx")
    
def parseAgent(agent,token):
    data = {}

    if 'os' in agent:
        if 'arch' in agent['os']:
            data['architecture'] = agent['os']['arch']
        else:
            data['architecture'] = "Unknown"

        if 'uname' in agent['os'] and 'version' in agent['os']:
            data['ComputerName'] = agent['os']['name']
            data['Operating_System'] = f"{agent['os']['uname']} {agent['os']['version']}" 
        else:
            data['ComputerName'] = "Unknown"
            data['Operating_System'] = "Unknown"
    else:
         data['architecture'] = "Unknown"
         data['ComputerName'] = "Unknown"
         data['Operating_System'] = "Unknown"
         
    data['AgentId'] = agent.get('id','Unknown')
    data['UserName'] = agent.get('name','Unknown')
    data['status'] = agent.get('status','Unknown')
    if 'registeredDate' in agent:

        data['registeredDate'] = print_time_ago(datetime.strptime(agent['dateAdd'],"%Y-%m-%dT%H:%M:%S+00:00").replace(tzinfo=pytz.utc)) 

    else:
        data['registeredDate'] = "Unknown"

    if 'lastKeepAlive' in agent:
         data['lastKeepAlive'] = print_time_ago(datetime.strptime(agent['lastKeepAlive'],"%Y-%m-%dT%H:%M:%S+00:00").replace(tzinfo=pytz.utc))
    else: 
        data['lastKeepAlive'] = 'unknown'

    store = getMoreData(data['AgentId'],token)
    data.update(store) 
    print(data)
    return data

def parseTime(dt):
    
    return datetime.strptime(dt,"%Y-%m-%dT%H:%M:%S+00:00").year 

def get_agents(token):  
    """Fetch the list of agents from Wazuh."""  
    url = f"{WAZUH_API_IP}/agents/?pretty=true"  
    headers = {  
        'Authorization': f'Bearer {token}',  
        'Content-Type': 'application/json'  
    }  
    
    response = requests.get(url, headers=headers,verify=False)  
    summary = {}
    if response.status_code == 200:  
        agents = response.json()  
        print("Agents list retrieved successfully:")
        # print(agents)
        print("**_**"*20)  
        summary = {}

        for index,agent in enumerate(agents['data']['affected_items']):
            index += 1
            saveToExcel(index,parseAgent(agent,token))  

    else:  
        print("Failed to retrieve agents:", response.json())  

       

def main():  
    token = authenticate()  
    get_agents(token) 

if __name__ == "__main__":  
    main()